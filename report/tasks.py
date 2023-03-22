import re
import datetime
from datetime import timedelta
import dateutil.parser
from django.db.models import Count, Sum, OuterRef, Subquery, Q, FloatField, F, Value as V
from django.db.models.functions import Coalesce
from celery import shared_task
from openpyxl import load_workbook
from os.path import basename

from agency.models import Agency
from main.models import Airline, Country, CommissionHistory
from report.models import ReportPeriod, ReportFile, Charges, AgencyDebitMemo, Transaction, Taxes, Remittance, \
    DailyCreditCardFile, ReprocessFile, Disbursement, CarrierDeductions, Deduction
from report.regex import file_header, period_header, summary_header, grand_total, \
    scope_combined, agency_details, transaction_details_exception, transaction_details, transaction_details_rtdn, \
    transaction_details_sub, newline, transaction_details_exception_left, transaction_details_exception_right, \
    transaction_details_exception_middle, c_file_header, c_card_details, c_transaction, date_header, cann, cr_header, \
    cr_ped, cr_airline, cr_seperater, transaction_details__no_transaction_amount, transaction_details__nr_code, \
    transaction_spdr, transaction_details_ec, transaction_details_guyana, transaction_detailsSA, transaction_detailsGY, \
    transaction_details_ec1, transaction_details__without_tax_and_fandc, transaction_detailsSA_less_data, \
    transaction_details_ec1, transaction_details_without_tax_to_pen, transaction_details_without_pen, \
    scope_international, transaction_details_till_tax_amount, \
    transaction_details_without_fandc_and_pen, transaction_details1, transaction_details_zimb, file_header_fr, \
    transaction_details_exception_left_ec, transaction_details_exception_right_ec, transaction_details_exception_right1, \
    transaction_details_sub_ec, grand_total_ca_single_line_ec, transaction_details_exception_left1

from report.utils import convert_amount, convert_transaction_date, convert_date, get_agency_no, get_float, change_format

from main.tasks import send_mail
from account.models import User


import logging



comined_totals = re.compile(r'^\s{0,1}COMBINED\s+TOTALS')
grand_total_ca = re.compile(r'^\s{0,1}GRAND TOTAL\s+(?P<type>\w{0,2})\s+(?P<transaction_amount>-?\d*,?\d*,?\d+\.\d{2})')
grand_total_ca_single_line = re.compile(
    r'^\s{0,1}GRAND TOTAL\s+(?P<type>\w{0,2})\s+(?P<transaction_amount>-?\d*,?\d*,?\d+\.\d{2})\s+(?P<fare_amount>-?\d*,?\d*,?\d+\.\d{2})\s+(?P<tax>-?\d*,?\d*,?\d+\.\d{2})\s+(?P<fandc>-?\d*,?\d*,?\d+\.\d{2})\s+(?P<pen>-?\d*,?\d*,?\d+\.\d{2})\s+(?P<cobl_amount>-?\d*,?\d*,?\d+\.\d{2})\s+(?P<std_comm>-?\d*,?\d*,?\d+\.\d{2})\s+(?P<supp_comm>-?\d*,?\d*,?\d+\.\d{2})\s+(?P<tax_on_comm>-?\d*,?\d*,?\d+\.\d{2})\s+(?P<balance>-?\d*,?\d*,?\d+\.\d{2})$')
grand_total_cc = re.compile(r'^\s+(?P<type>\w{0,2})\s+(?P<transaction_amount>-?\d*,?\d*,?\d+\.\d{2})')
values_totals = re.compile(
    r'^\s+\*{2}\s+(?P<transaction_amount>-?\d*,?\d*,?\d+\.\d{2})\s+(?P<fare_amount>-?\d*,?\d*,?\d+\.\d{2})\s+(?P<tax>-?\d*,?\d*,?\d+\.\d{2})\s+(?P<fandc>-?\d*,?\d*,?\d+\.\d{2})\s+(?P<pen>-?\d*,?\d*,?\d+\.\d{2})\s+(?P<cobl_amount>-?\d*,?\d*,?\d+\.\d{2})\s+(?P<std_comm>-?\d*,?\d*,?\d+\.\d{2})\s+(?P<supp_comm>-?\d*,?\d*,?\d+\.\d{2})\s+(?P<tax_on_comm>-?\d*,?\d*,?\d+\.\d{2})\s+(?P<balance>-?\d*,?\d*,?\d+\.\d{2})')
credit_memos = re.compile(
    r'^\s{0,2}CREDIT MEMOS\s+(?P<type>\w{0,2})\s+(?P<transaction_amount>-?\d*,?\d*,?\d+\.\d{2})\s+(?P<fare_amount>-?\d*,?\d*,?\d+\.\d{2})\s+(?P<tax_amount>-?\d*,?\d*,?\d+\.\d{2})\s+(?P<fandc>-?\d*,?\d*,?\d+\.\d{2})\s+(?P<pen>-?\d*,?\d*,?\d+\.\d{2})\s+(?P<cobl_amount>-?\d*,?\d*,?\d+\.\d{2})\s+(?P<std_comm>-?\d*,?\d*,?\d+\.\d{2})\s+(?P<supp_comm>-?\d*,?\d*,?\d+\.\d{2})\s+(?P<tax_on_comm>-?\d*,?\d*,?\d+\.\d{2})\s+(?P<balance>-?\d*,?\d*,?\d+\.\d{2})')

no_data_in_credit_file = re.compile(r'\s*NO DATA TO REPORT THIS DATE')

from celery import task


def datetime_range(start=None, end=None):
    span = end - start
    for i in range(span.days + 1):
        yield start + timedelta(days=i)


@task
def delete_obj(obj_id):
    try:
        ReportFile.objects.get(id=obj_id).delete()
    except Exception:
        pass


def create_adm_or_acm(amount, comment, transaction, allowed_commission_amount, is_acm=False):
    adm, created = AgencyDebitMemo.objects.update_or_create(transaction=transaction, defaults={
        "amount": amount,
        "comment": comment,
        "is_acm": is_acm,
        "allowed_commission_amount": allowed_commission_amount
    })
    return True


def create_charges_and_taxes(transaction_details_values, transaction, countryname=""):
    tax_amount = transaction_details_values.get('tax_amount', None)
    fandc_amount = transaction_details_values.get('fandc_amount', None)
    pen_amount = transaction_details_values.get('pen', None)
    if countryname and countryname.lower() in ["netherlands",
                                               "ecuador"]:  # currency format is different for these countries from others.
        tax_amount = change_format(tax_amount)
        fandc_amount = change_format(fandc_amount)
        pen_amount = change_format(pen_amount)
    else:
        tax_amount = convert_amount(tax_amount)
        fandc_amount = convert_amount(fandc_amount)
        pen_amount = convert_amount(pen_amount)
    tax_type = transaction_details_values.get('tax_type', None)
    fandc_type = transaction_details_values.get('fandc_type', None)
    pen_type = transaction_details_values.get('pen_type', None)
    if tax_amount or tax_type:
        if tax_type in ['CP', 'MF', 'YQ', 'YR']:

            charge = Charges.objects.create(amount=tax_amount, type=tax_type,
                                            transaction=transaction)
        else:
            tax = Taxes.objects.create(amount=tax_amount, type=tax_type, transaction=transaction)

    if fandc_amount or fandc_type:
        if countryname and countryname.lower() == "zimbabwe":  # in Zimbabwe report files, f&c type can ve empty.
            if fandc_type in ['CP', 'MF', 'YQ', 'YR'] or fandc_type is None:
                charge = Charges.objects.create(amount=fandc_amount, type=fandc_type,
                                                transaction=transaction)
        elif fandc_type and fandc_type not in ['CP', 'MF', 'YQ', 'YR']:
            tax = Taxes.objects.create(amount=fandc_amount, type=fandc_type, transaction=transaction)
        elif fandc_type in ['CP', 'MF', 'YQ', 'YR']:

            charge = Charges.objects.create(amount=fandc_amount, type=fandc_type,
                                            transaction=transaction)

    if pen_amount or pen_type:
        if pen_type in ['CP', 'MF', 'YQ', 'YR']:
            charge = Charges.objects.create(amount=pen_amount, type=pen_type,
                                            transaction=transaction)
        elif pen_type and pen_type not in ['CP', 'MF', 'YQ', 'YR']:
            tax = Taxes.objects.create(amount=pen_amount, type=pen_type, transaction=transaction)
    return True


def create_transaction(agency, rf, transaction_details_values, countryname=None):
    fop = transaction_details_values.get('fop')
    stat = transaction_details_values.get('stat')
    cpui = transaction_details_values.get('cpui')
    is_sale = True

    cc = None
    ca = None
    ep = None

    if fop in ['CC', 'CA', 'EX' , 'EP']:
        pass
    elif transaction_details_values.get('transaction_type') in ['RFND', 'ACMA', 'SPCR', 'ADMA']:
        fop = transaction_details_values.get('stat')
        stat = transaction_details_values.get('cpui')
        cpui = None
    else:
        stat = fop
        fop = None

    if fop:
        if fop == 'CC':
            if transaction_details_values.get('CC', None):
                cc = transaction_details_values.get('CC', None)
            else:
                cc = transaction_details_values.get('transaction_amount')
            ca = transaction_details_values.get('CA', None)
        elif fop == 'CA':
            if transaction_details_values.get('CA', None):
                ca = transaction_details_values.get('CA', None)
            else:
                ca = transaction_details_values.get('transaction_amount')
            cc = transaction_details_values.get('CC', None)
        elif fop == 'EP':
            if transaction_details_values.get('EP', None): 
                ep = transaction_details_values.get('EP', None)
            else:  
                ep = transaction_details_values.get('transaction_amount')
    
    else:
        cc = None
        ca = None
        ep = None
    if countryname and countryname.lower() in ["netherlands",
                                               "ecuador"]:  # currency format is different for these countries from others.
        std_comm_rate = change_format(transaction_details_values.get('std_comm_rate'))
        std_comm_amount = change_format(transaction_details_values.get('std_comm_amount'))
        sup_comm_rate = change_format(transaction_details_values.get('sup_comm_rate'))
        sup_comm_amount = change_format(transaction_details_values.get('sup_comm_amount'))
        fare_amount = change_format(transaction_details_values.get('fare_amount'))
        cobl_amount = change_format(transaction_details_values.get('cobl_amount'))
        transaction_amount = change_format(transaction_details_values.get('transaction_amount'))
        ca = change_format(ca)
        cc = change_format(cc)
        ep = change_formate(ep)
        balance = float(change_format(transaction_details_values.get('balance')))
        tax_on_comm = float(change_format(transaction_details_values.get('tax_on_comm')))


    else:
        std_comm_rate = convert_amount(transaction_details_values.get('std_comm_rate'))
        std_comm_amount = float(convert_amount(transaction_details_values.get('std_comm_amount')))
        sup_comm_rate = convert_amount(transaction_details_values.get('sup_comm_rate'))
        sup_comm_amount = convert_amount(transaction_details_values.get('sup_comm_amount'))
        fare_amount = float(convert_amount(transaction_details_values.get('fare_amount')))
        cobl_amount = convert_amount(transaction_details_values.get('cobl_amount'))
        transaction_amount = convert_amount(transaction_details_values.get('transaction_amount'))
        ca = convert_amount(ca)
        cc = convert_amount(cc)
        ep = convert_amount(ep)
        balance = float(convert_amount(transaction_details_values.get('balance')))
        tax_on_comm = float(convert_amount(transaction_details_values.get('tax_on_comm')))
        cobl_amount = float(convert_amount(str(cobl_amount)))

    if fop == 'CC':
        is_sale = False
    transaction, created_at = Transaction.objects.get_or_create(
        agency=agency,
        report=rf,
        ticket_no=transaction_details_values.get('ticket_no'),
        transaction_type=transaction_details_values.get('transaction_type'),
        issue_date=convert_transaction_date(transaction_details_values.get('issue_date')),
    )
    transaction.transaction_amount = transaction_amount
    transaction.fare_amount = float(fare_amount)
    transaction.cobl_amount = cobl_amount
    transaction.std_comm_rate = float(std_comm_rate)
    transaction.std_comm_amount = float(std_comm_amount)
    transaction.sup_comm_rate = float(sup_comm_rate)
    transaction.sup_comm_amount = float(sup_comm_amount)
    transaction.tax_on_comm = tax_on_comm
    transaction.balance = balance
    transaction.cpui = cpui
    transaction.stat = stat
    transaction.fop = fop
    transaction.cc = cc
    transaction.ca = ca
    transaction.ep = ep
    transaction.is_sale = is_sale
    transaction.save()
    return transaction, created_at


def process_billing_details(text_file=None, request=None, report_file=None):
    """Parse billing details file"""

    credit_memos_values = None
    grand_total_ca_values = None
    grand_total_cc_values = None
    values_totals_values = None
    grand_total_found = False

    if report_file:
        filedata = open(report_file.file.path, 'r', encoding="utf-8")
    else:
        filedata = open(text_file, 'r', encoding="utf-8")
    reader = enumerate(filedata.readlines())
    number, line = next(reader, None)
    m = re.match(file_header, line)
    if m is None:
        m = re.match(file_header_fr, line)
    if report_file:
        country = report_file.country
    else:
        if request:
            # country = Country.objects.get(id=request.session.get('country'))
            if request.POST.get("from_scheduler") is not None:
                country = Country.objects.get(id=request.POST.get("countrycode"))
            else:
                country = Country.objects.get(id=request.session.get('country'))
        else:
            country = None
    if m:
        header_values = m.groupdict()
        try:
            airline = Airline.objects.get(code=header_values.get('code'), country=country)
            allowed_commission_rate = 0.00

        except:
            return ("It seems like there is no airline with this 3 digit code : {}".format(header_values.get('code')))
    else:

        return ("File does not seem to have the right format")

    if not report_file:
        number, line = next(reader, None)

        m = re.match(period_header, line)
        period_values = None
        while not m:
            number, line = next(reader, None)

            m = re.match(period_header, line)
        else:
            period_values = m.groupdict()
        if period_values:
            try:
                start_date = convert_date(period_values.get('start'))
                end_date = convert_date(period_values.get('end'))
                # report_period = ReportPeriod.from_date(end_date)
                report_period = ReportPeriod.objects.get(ped=end_date, country=country)
                # ojb= ReportFile.objects.filter(report_period=report_period, airline=airline, country=country, 
                #         ).delete()
                rf, created = ReportFile.objects.update_or_create(
                    report_period=report_period, airline=airline, country=country, defaults={
                        'ref_no': period_values.get('ref_no').strip(),
                        'file': text_file.split('media/')[-1],
                    })
                commission_history = CommissionHistory.objects.filter(airline=airline, type='M',
                                                                      from_date__lte=end_date)
                if commission_history.exists():
                    allowed_commission_rate = commission_history.order_by('-from_date').first().rate
            except:
                return ("Billing Period date invalid format/ PED is not found")
        else:
            return ("Billing Period data not found")
    else:
        rf = report_file
    try:

        # scope combined section
        number, line = next(reader, None)

        m = re.match(scope_combined, line)
        n = re.match(scope_international, line)
        is_scope_combined_found = False
        while not m and not n:

            number, line = next(reader, None)

            m = re.match(scope_combined, line)
            n = re.match(scope_international, line)
        else:
            is_scope_combined_found = True

    except:
        return False
    if is_scope_combined_found:

        transaction = False
        try:
            while reader:
                number, line = next(reader, None)
                m = re.match(agency_details, line)

                while not m:
                    number, line = next(reader, None)
                    if (re.match(newline, line)):
                        continue
                    m = re.match(agency_details, line)
                    transaction_details_values = None

                    if country.name == "Ecuador" and re.match(transaction_details1, line) is None and (
                        ",090505" in line or ",092302" in line or "Ecuador ,0910164" in line):  # remove hidden values present between the values
                        line_splitted = None
                        if ",090505" in line:
                            line_splitted = line.split(",090505")[0]
                        elif ",092302" in line:
                            line_splitted = line.split(",092302")[0]
                        elif ",0910164" in line:
                            line_splitted = line.split("Ecuador ,0910164")[0]

                        if report_file:
                            filedata1 = open(report_file.file.path, 'r', encoding="utf-8")
                        else:
                            filedata1 = open(text_file, 'r', encoding="utf-8")
                        reader_1 = enumerate(filedata1.readlines())
                        reader_dict = dict(list(reader_1))
                        data = reader_dict[number + 1]
                        converted_str = ""
                        if line_splitted:
                            for strng in line_splitted.split(" "):
                                if strng:
                                    if converted_str:
                                        converted_str += "      " + strng
                                    else:
                                        converted_str += strng
                            cnvrtd_line = converted_str + "           " + data.lstrip()
                            line = cnvrtd_line
                    mcann = re.match(cann, line)
                    if airline.name.lower() == "uganda airlines":  # remove unnecessary * from values

                        if " *" in line:
                            line = line.replace(" *", "")

                    if mcann:

                        transaction_details_values = mcann.groupdict()
                        transaction, created_t = create_transaction(agency, rf, transaction_details_values)
                    elif re.match(transaction_spdr, line):
                        transaction_details_values = re.match(transaction_spdr, line).groupdict()
                        transaction_details_values['transaction_type'] = 'SPDR'
                        transaction, created_t = create_transaction(agency, rf, transaction_details_values)
                    elif country.name.lower() not in ["zimbabwe"] and re.match(transaction_details, line):

                        if re.match(transaction_details, line):

                            transaction_details_values = re.match(transaction_details, line).groupdict()
                            transaction, created_t = create_transaction(agency, rf, transaction_details_values)
                            if not created_t:
                                transaction.taxes_set.all().delete()
                                transaction.charges_set.all().delete()
                                transaction.agencydebitmemo_set.all().delete()
                            create_charges_and_taxes(transaction_details_values, transaction, countryname=country.name)

                    elif country.name.lower() in ["zimbabwe"] and re.match(transaction_details_zimb, line):

                        if re.match(transaction_details_zimb, line):

                            transaction_details_values = re.match(transaction_details_zimb, line).groupdict()
                            transaction, created_t = create_transaction(agency, rf, transaction_details_values)
                            if not created_t:
                                transaction.taxes_set.all().delete()
                                transaction.charges_set.all().delete()
                                transaction.agencydebitmemo_set.all().delete()
                            create_charges_and_taxes(transaction_details_values, transaction, countryname=country.name)

                    elif re.match(transaction_details1, line):

                        if re.match(transaction_details1, line):

                            transaction_details_values = re.match(transaction_details1, line).groupdict()
                            transaction, created_t = create_transaction(agency, rf, transaction_details_values,
                                                                        countryname=country.name)
                            if not created_t:
                                transaction.taxes_set.all().delete()
                                transaction.charges_set.all().delete()
                                transaction.agencydebitmemo_set.all().delete()
                            create_charges_and_taxes(transaction_details_values, transaction, countryname=country.name)
                    elif country.name == "Ecuador" and re.match(transaction_details_ec1, line):

                        if re.match(transaction_details_ec1, line):

                            transaction_details_values = re.match(transaction_details_ec1, line).groupdict()
                            transaction, created_t = create_transaction(agency, rf, transaction_details_values,
                                                                        countryname=country.name)
                            if not created_t:
                                transaction.taxes_set.all().delete()
                                transaction.charges_set.all().delete()
                                transaction.agencydebitmemo_set.all().delete()
                            create_charges_and_taxes(transaction_details_values, transaction, countryname=country.name)

                    elif re.match(transaction_details_without_pen, line):
                        if re.match(transaction_details_without_pen, line):

                            transaction_details_values = re.match(transaction_details_without_pen, line).groupdict()
                            transaction, created_t = create_transaction(agency, rf, transaction_details_values)
                            if not created_t:
                                transaction.taxes_set.all().delete()
                                transaction.charges_set.all().delete()
                                transaction.agencydebitmemo_set.all().delete()
                            create_charges_and_taxes(transaction_details_values, transaction)
                    elif re.match(transaction_details_without_tax_to_pen, line):
                        if re.match(transaction_details_without_tax_to_pen, line):

                            transaction_details_values = re.match(transaction_details_without_tax_to_pen,
                                                                  line).groupdict()
                            transaction, created_t = create_transaction(agency, rf, transaction_details_values)
                            if not created_t:
                                transaction.taxes_set.all().delete()
                                transaction.charges_set.all().delete()
                                transaction.agencydebitmemo_set.all().delete()
                            create_charges_and_taxes(transaction_details_values, transaction)
                    elif country.name.lower() not in ["canada", 'united states'] and re.match(
                        transaction_details_till_tax_amount, line):
                        if re.match(transaction_details_till_tax_amount, line):

                            transaction_details_values = re.match(transaction_details_till_tax_amount, line).groupdict()

                            transaction, created_t = create_transaction(agency, rf, transaction_details_values)
                            if not created_t:
                                transaction.taxes_set.all().delete()
                                transaction.charges_set.all().delete()
                                transaction.agencydebitmemo_set.all().delete()
                            create_charges_and_taxes(transaction_details_values, transaction)

                    elif country.name.lower() not in ["canada", 'united states'] and re.match(
                        transaction_details_without_fandc_and_pen, line):
                        if re.match(transaction_details_without_fandc_and_pen, line):

                            transaction_details_values = re.match(transaction_details_without_fandc_and_pen,
                                                                  line).groupdict()
                            transaction, created_t = create_transaction(agency, rf, transaction_details_values)

                            if not created_t:
                                transaction.taxes_set.all().delete()
                                transaction.charges_set.all().delete()
                                transaction.agencydebitmemo_set.all().delete()
                            create_charges_and_taxes(transaction_details_values, transaction)
                    elif re.match(transaction_details__no_transaction_amount, line):
                        transaction_details_values = re.match(transaction_details__no_transaction_amount,
                                                              line).groupdict()

                        transaction, created_t = create_transaction(agency, rf, transaction_details_values)
                    elif re.match(transaction_details__nr_code, line):
                        transaction_details_values = re.match(transaction_details__nr_code, line).groupdict()
                        transaction, created_t = create_transaction(agency, rf, transaction_details_values)

                        if not created_t:
                            transaction.taxes_set.all().delete()
                            transaction.charges_set.all().delete()
                            transaction.agencydebitmemo_set.all().delete()
                        create_charges_and_taxes(transaction_details_values, transaction)
                    elif re.match(transaction_details_exception_left, line):
                        transaction_details_values = re.match(transaction_details_exception_left, line).groupdict()

                        fop = transaction_details_values.get('fop', None)
                        if fop:
                            transaction_details_values.update({transaction_details_values.get(
                                'fop'): transaction_details_values.get('transaction_amount')})

                        cu_reader = reader
                        if report_file:
                            filedata1 = open(report_file.file.path, 'r', encoding="utf-8")
                        else:
                            filedata1 = open(text_file, 'r', encoding="utf-8")
                        reader_1 = enumerate(filedata1.readlines())
                        number, line = next(reader_1, None)
                        ml = re.match(transaction_details_exception_right, line)
                        try:
                            while not ml:
                                number, line = next(reader_1, None)
                                mn = re.match(transaction_details_exception_middle, line)
                                if mn:
                                    transaction_details_exception_middle_values = mn.groupdict()
                                    if transaction_details_exception_middle_values.get('fop') in ['CC', 'CA', 'EP']:
                                        transaction_details_values.update(
                                            {transaction_details_exception_middle_values.get(
                                                'fop'): transaction_details_exception_middle_values.get(
                                                'transaction_amount')})

                                ml = re.match(transaction_details_exception_right, line)
                            else:
                                transaction_details_exception_right_values = ml.groupdict()
                                transaction_details_values.update(transaction_details_exception_right_values)

                        except:
                            number, line = next(cu_reader, None)

                    if country.name == "Ecuador":
                        mj = re.match(transaction_details_sub_ec, line)

                    else:
                        mj = re.match(transaction_details_sub, line)
                    if mj and transaction:

                        transaction_details_sub_values = mj.groupdict()
                        try:
                            create_charges_and_taxes(transaction_details_sub_values, transaction,
                                                     countryname=country.name)
                        except Exception as e:
                            print("Error Occured : ", e)
                    mk = re.match(transaction_details_rtdn, line)
                    try:

                        if airline.name.lower() != "uganda airlines":
                            transaction, created_t = create_transaction(agency, rf, transaction_details_values,
                                                                        countryname=country.name)

                    except:
                        pass
                    if mk:
                        transaction_details_rtdn_values = mk.groupdict()
                        is_sale = True
                        if transaction_details_rtdn_values.get('fop', None) == 'CC':
                            is_sale = False

                        transaction_rtdn, created_tr = Transaction.objects.update_or_create(
                            agency=agency,
                            report=rf,
                            ticket_no=transaction_details_rtdn_values.get('ticket_no'),
                            transaction_type=transaction_details_rtdn_values.get('transaction_type'),
                            defaults={
                                'transaction_amount': convert_amount(
                                    transaction_details_rtdn_values.get('transaction_amount')),
                                'fop': transaction_details_rtdn_values.get('fop', None),
                                'is_sale': is_sale
                            },
                        )

                    mct = re.match(comined_totals, line)

                    if mct and not grand_total_found:

                        grand_total_found = True

                        number, line = next(reader, None)

                        ma = re.match(grand_total_ca, line)
                        while not ma:
                            number, line = next(reader, None)
                            ma = re.match(grand_total_ca, line)
                            mcm = re.match(credit_memos, line)
                            if mcm:
                                credit_memos_values = mcm.groupdict()
                        else:
                            if country.name == "Ecuador":
                                msl = re.match(grand_total_ca_single_line_ec, line)
                            else:
                                msl = re.match(grand_total_ca_single_line, line)

                            if msl:
                                grand_total_ca_values_full = msl.groupdict()
                                grand_total_ca_values = {}
                                grand_total_ca_values['type'] = grand_total_ca_values_full.pop('type', 'CC')
                                grand_total_ca_values['transaction_amount'] = grand_total_ca_values_full.get(
                                    'transaction_amount', 0.00)
                                values_totals_values = grand_total_ca_values_full
                            else:
                                grand_total_ca_values = ma.groupdict()
                            number, line = next(reader, None)

                            mc = re.match(grand_total_cc, line)
                            if mc:
                                grand_total_cc_values = mc.groupdict()
                            number, line = next(reader, None)
                            mv = re.match(values_totals, line)
                            if mv:
                                values_totals_values = mv.groupdict()


                else:
                    agency_details_values = m.groupdict()
                    trade_name = agency_details_values.get('trade_name')
                    if (trade_name.find('GST') != -1):
                        trade_name = trade_name[0:trade_name.find('GST')]
                    if (trade_name.find('HST') != -1):
                        trade_name = trade_name[0:trade_name.find('HST')]

                    trade_name = trade_name.strip()
                    agency, created = Agency.objects.get_or_create(
                        agency_no=get_agency_no(agency_details_values.get('agency_no')), country=country,
                        defaults={
                            'trade_name': trade_name
                        },
                    )

        except TypeError:
            pass

        data = {}

        if grand_total_ca_values:
            data[grand_total_ca_values.get("type", "cc").lower()] = grand_total_ca_values.get("transaction_amount",
                                                                                              None)
        if grand_total_cc_values:
            data[grand_total_cc_values.get("type", "ca").lower()] = grand_total_cc_values.get("transaction_amount",
                                                                                              None)
        if values_totals_values:
            data.update(values_totals_values)

        def get_data_from_transaction_table(data=None):
            trans_data = Transaction.objects.exclude(transaction_type__startswith='SP',
                                                     agency__agency_no='6999001').filter(report=rf).aggregate(
                transaction_amount=Sum('transaction_amount'),
                fare_amount=Sum('fare_amount'),
                std_comm=Sum('std_comm_amount'),
                cobl_amount=Sum('cobl_amount'),
                supp_comm=Sum('sup_comm_amount'),
                tax_on_comm=Sum('tax_on_comm'),
                balance=Sum('balance'),
                cc=Sum('cc'),
                ca=Sum('ca')
            )
            tax_tot = Taxes.objects.filter(transaction__report=rf).aggregate(total_tax=Sum('amount')).get('total_tax')
            tot_fandc = Charges.objects.filter(transaction__report=rf, type__in=['YQ', 'YR']).aggregate(
                total_charges=Sum('amount')).get('total_charges')
            tot_pen = Charges.objects.filter(transaction__report=rf, type__in=['CP', 'MF']).aggregate(
                total_charges=Sum('amount')).get('total_charges')

            data.update(trans_data)
            data['tax'] = tax_tot if tax_tot else 0.00
            data['pen'] = tot_pen if tot_pen else 0.00
            data['fandc'] = tot_fandc if tot_fandc else 0.00
            return data

        if not data:
            data = get_data_from_transaction_table(data)

        if credit_memos_values:
            data["acms"] = credit_memos_values.get("transaction_amount", None)

        check_data = get_data_from_transaction_table(data={})

        try:
            for i in values_totals_values:

                if country.name == "Ecuador":
                    values_totals_values[i] = float(change_format(values_totals_values[i]))
                else:
                    values_totals_values[i] = float(convert_amount(values_totals_values[i]))

            err_flag = False
            err_vals = []
            key_arr = ["transaction_amount", "fare_amount", "tax", "fandc", "pen", "cobl_amount", "std_comm",
                       "supp_comm", "tax_on_comm", "balance"]
            for key in key_arr:
                if float(values_totals_values.get(key)) != float(num_quantize(check_data.get(key))):
                    err_flag = True
                    err_dict = {
                        'key': key,
                        'table_val': check_data.get(key),
                        'report_val': values_totals_values.get(key)
                    }
                    err_vals.append(err_dict)
            # else:

        except:
            pass

        if data:
            data = {k: str(data[k]).replace(',', '') for k in data if k}
            ReportFile.objects.filter(pk=rf.pk).update(**data)

    else:
        return ("Error occured while parsing file.")


credit_file_grand_total = re.compile(
    r'\s{0,2}GRAND TOTAL\s+(?P<tax>-?\d*,?\d*,?\d+\.\d{2})\s+(?P<cc>-?\d*,?\d*,?\d+\.\d{2})')


def process_card_details(text_file, request=None):
    """Parse card details file"""

    filedata = open(text_file, 'r', encoding="utf-8")
    grand_total_values = {}
    if request:
        if request.POST.get("from_scheduler") is not None:
            country = Country.objects.get(id=request.POST.get("countrycode"))
        else:
            country = Country.objects.get(id=request.session.get('country'))
    else:
        country = None

    reader = enumerate(filedata.readlines())
    number, line = next(reader, None)
    m = re.match(c_file_header, line)

    if m:
        header_values = m.groupdict()
        try:
            airline = Airline.objects.get(code=header_values.get('code'), country=country)
        except  :
            return ("It seems like there is no airline with this 3 digit code.")
    else:
        return ("File does not seem to have the right format")

    number, line = next(reader, None)
    m = re.match(date_header, line)
    date_values = None
    while not m:
        number, line = next(reader, None)
        m = re.match(date_header, line)
    else:
        date_values = m.groupdict()
    if date_values:
        from_date = convert_date(date_values.get('date')) - datetime.timedelta(1)
        try:
            report_period = ReportPeriod.objects.filter(ped__gte=from_date, country=country).order_by('ped').first()
        except  :
            return 'PED is not found'

        if not ReportFile.objects.filter(airline=airline, report_period=report_period).exists():
            return ("It seems like there is no billing details file uploaded for this week.")
        cf, created = DailyCreditCardFile.objects.update_or_create(
            date=convert_date(date_values.get('date')), airline=airline, defaults={"from_date": from_date})

    # summary section
    try:
        while reader:
            number, line = next(reader, None)
            m = re.match(c_card_details, line)
            while not m:
                # skip new lines
                if (re.match(newline, line)):
                    number, line = next(reader, None)
                    continue

                if (re.match(no_data_in_credit_file, line)):
                    break

                # grand  total
                m_grand = re.match(credit_file_grand_total, line)
                if m_grand:
                    grand_total_values = m_grand.groupdict()

                # transaction details
                mi = re.match(c_transaction, line)
                while mi:
                    c_transaction_values = mi.groupdict()
                    agency = get_agency_no(c_transaction_values.get('agency_no'))
                    issue_date = convert_transaction_date(c_transaction_values.get('issue_date'))
                    card_type = c_card_details_values.get('card')
                    try:
                        trans = Transaction.objects.get(agency__agency_no=agency,
                                                        transaction_type=c_transaction_values.get(
                                                            'transaction_type'),
                                                        ticket_no=c_transaction_values.get('ticket_no'),
                                                        issue_date=issue_date)
                        trans.card_type = card_type
                        trans.is_sale = True
                        trans.save()

                        airline = trans.report.airline
                        allowed_commission_rate = 0.00
                        commission_history = CommissionHistory.objects.filter(airline=airline, type='M',
                                                                              from_date__lte=report_period.ped)
                        if commission_history.exists():
                            allowed_commission_rate = commission_history.order_by('-from_date').first().rate

                        allowed_commission_amount = (trans.cobl_amount * allowed_commission_rate) / 100
                        if c_transaction_values.get('transaction_type', None) == 'TKTT':

                            if card_type.lower().strip() == 'american express' and airline.accepts_AMEX:
                                try:
                                    AgencyDebitMemo.objects.get(transaction=trans).delete()
                                except  :
                                    pass

                            elif card_type.lower().strip() == 'mastercard' and airline.accepts_MC:
                                try:
                                    AgencyDebitMemo.objects.get(transaction=trans).delete()
                                except  :
                                    pass

                            elif card_type.lower().strip() == 'visa international' and airline.accepts_VI:
                                try:
                                    AgencyDebitMemo.objects.get(transaction=trans).delete()
                                except  :
                                    pass



                    except Transaction.DoesNotExist:
                        pass
                    number, line = next(reader, None)
                    mi = re.match(c_transaction, line)

                number, line = next(reader, None)
                m = re.match(c_card_details, line)
            else:
                c_card_details_values = m.groupdict()
    except TypeError:
        cf.grand_total = convert_amount(grand_total_values.get('cc', '0.00'))
        cf.save()
        start_date = report_period.from_date
        end_date = report_period.ped

        dates_list = list(datetime_range(start=start_date, end=end_date))
        count_of_daily = DailyCreditCardFile.objects.filter(airline=airline, from_date__in=dates_list)
        if count_of_daily.count() == len(dates_list) and count_of_daily.count() != 0:
            # check sum
            total_cc = round(
                count_of_daily.aggregate(total_cc=Coalesce(Sum('grand_total'), V(0))).get('total_cc', 0.00), 2)
            trans_total_cc = round(
                Transaction.objects.exclude(transaction_type__startswith='SP', agency__agency_no='6999001').filter(
                    report__report_period=report_period, report__airline=airline, report__country=country).aggregate(
                    total_cc=Coalesce(Sum('cc'), V(0))).get('total_cc', 0.00), 2)
            if total_cc != trans_total_cc:
                # send mail here
                try:
                    rf = ReportFile.objects.filter(airline=airline, report_period=report_period).first()
                    file_name = rf.file.name.split('/')[-1].split('.')[0]
                except  :
                    file_name = ''
                context = {
                    'request': request,
                    'error': 'Credit Card Value Mismatch',
                    'total_cc': total_cc,
                    'trans_total_cc': trans_total_cc,
                    'file_name': file_name,
                }
                admin_emails = User.objects.filter(is_superuser=True).values_list('email', flat=True)
                try:
                    send_mail("Credit Card Value Mismatch.", "email/cc-mismatch-issue-email.html", context,
                              admin_emails, from_email='assda@assda.com')
                except Exception as e:
                    print("Error Occured : ", e)


####ARC
currency = '\d*,?\d*,?\d*\.?\d+-?'

_process_funcs = []


def register(regexp):
    def wrapped(func):
        _process_funcs.append(func)
        func.regexp = regexp
        return func

    return wrapped


class ProcessData(object):
    separator = re.compile('^-{132}$')

    transaction_header = re.compile(
        '^(\d{2}) (\d{5})-(\d).+?([\w\.]+)[ ]+'  # ag_no+title
        'BOX/BATCH CODE - ([\d-]+)[ ]+'  # batch code
        'AGENT RPT PERIOD (\d{2}/\d{2}/\d{2})$')  # date

    ticket_header = re.compile(
        '^(\d\d\d\d)?[ ]+'  # card no
        '(\d{10})[ ]+'  # tkt_no
        '[A-Z]?[ ]+'
        '[\dA][ ]+'
        '(\d{2}/\d{2})? (X)?[ ]+'  # date (unused)
        '(' + currency + ')[ ]+'  # fare incl tax
                         '(' + currency + ')[ ]+'  # fare ex tax
                                          '(' + currency + ')[ ]+'  # commission
                                                           '(\d*\.\d)D[ ]+'  # rate
                                                           '(' + currency + ')[ ]+'  # tax
                                                                            '\d\dD[ ]+'
                                                                            '([A-Z0-9]{2})[ ]+'  # tax type
        + currency + '[ ]*\d*$'  # not sure what this value is
    )

    tax_id = re.compile(
        '^[\w.\- ]+[ ]'
        '(' + currency + ')[ ]+'  # tax
                         '\d\dD[ ]+'
                         '([A-Z0-9]{2})$'  # tax type
    )

    even_exchange_transaction = re.compile(
        '^[ ]+(\d{10})[ ]+'  # ticket number
        '\d[ ]+EVEN EXCH[ ]+'
        '(' + currency + ')[ ]+'  # charge (should be 0)
                         '(' + currency + ')[ ]+'  # commission
                                          '\d+\.\dD[ ]+'
                                          '(' + currency + ')[ ]+'  # tax
                                                           '\d\dD ([0-9A-Z]+)[ ]+'  # tax type
        + currency + '[ ]*'
                     '\d*$'
    )

    refund_or_exchange_transaction = re.compile(
        '^(\d?\d?\d?\d?)[ ]+'  # cardcode
        '(\d+)[ ]+'  # txt_no
        '[C|O| ]C?R?N? A?O?[ ]+'  # card refund or cash? refund
        '\d?[ ]+'
        '(\d*)/?(\d*)[ ]+'  # date
        '(X?)[ ]+'  # X tells if international sale
        '(' + currency + ')[ ]+'  # total ref
                         '(' + currency + ')[ ]+'  # re tax ref
                                          '(' + currency + ')[ ]+'  # commission ref
                                                           '\d+\.\dD[ ]+'
                                                           '(' + currency + ')[ ]+'  # tax ref
                                                                            '\d\dD[ ]+'
                                                                            '([0-9A-Z]{2})[ ]+'  # tax type
        + currency + '[ ]*'
                     '\d*$'
    )

    exchange_transaction = re.compile(
        '^(\d?\d?\d?\d?)[ ]+'  # card code
        'EXCH-[ ]+'
        '\d+[\d ]*-?\d*-(\d+)[ ]*'  # ticket number
        '(\d*,?\d*\.?\d*-?)[ ]*'  # penalty
        '(\*?P?)[ ]*\d*$'  # test
    )

    refund_transaction = re.compile(
        '^(\d?\d?\d?\d?)[ ]+'  # card code
        '([R|E][F|X][N|C][D|H])-[ ]+'  # refund or exchange identifier
        '\d+[\d ]*-?\d*-(\d+)[ ]*'  # ticket number
        '(\d*,?\d*\.?\d*-?)[ ]*'  # penalty
        '(\*?P?)[ ]*'  # test
        '\d*$'
    )

    credit_debit_note_transaction = re.compile(
        '[ ]+'
        '(\d+)[ ]+'  # identified number
        'ADJ\.[ ]+'
        '(' + currency + ')[ ]*'  # amount
                         '\d*$'  # TODO: Unknown value
    )

    AAD_transaction = re.compile(
        '[ ]+'
        '(897\d+)[ ]+'  # identified number
        'AAD[ ]+'
        '\d{3}[ ]+'  # no idea
        '(\d+)[ ]+'  # ticket number
        'ORIG RPTED[ ]+'  # could be anything
        '[0-9\-]{9}[ ]+'
        '(' + currency + ')[ ]+'
                         '\d+$'
    )

    void_header = re.compile(
        '^(\d{2}) (\d{5})-(\d)[\w &/-]+'  # agency number
        'REPORT REFERENCE NUMBER[ ]+'
        '([\d-]+)[ ]+'  # ref number
        'AGENT RPT PERIOD - '
        '(\d{2}/\d{2}/\d{2})$'  # date
    )

    void_ticket = re.compile(
        '^[VM][ ]+'
        '(\d?\d?\d?\d?)'  # card number
        '[ ]+'  # wild card for manual void
        '(\d+)[ ]\d?[ ]+'  # ticket number
        '(?:\d{2}/\d{2})[ ]+'  # date (unused)
        '(' + currency + ')[ ]+'  # total
        + currency + '$'
    )

    can_void_ticket = re.compile(
        '^C[ ]+'
        '\*?\*?\*?\*?'  # card number
        '[ ]+'  # wild card for manual void
        '(\d+)[ ]\*[ ]+'  # ticket number
        '(?:\d{2}/\d{2})[ ]+'  # date (unused)
        + currency + '$'
    )

    transaction_summary = re.compile(
        '^[ ]*\d+\/[ ]*\d+\/[ ]*\d+[ ]+'
        '(' + currency + ')[ ]+'  # credit
                         '(' + currency + ')[ ]+'  # cash
                                          '(' + currency + ')[ ]+'  # total
                                                           '(' + currency + ')[ ]+'  # fare
                                                                            '(' + currency + ')[ ]+'  # commission
                                                                                             '(' + currency + ')[ ]+'  # tax
                                                                                                              '(' + currency + ')[ ]+'  # adjust
                                                                                                                               '(' + currency + ')[ ]+'  # ccrn_comm
                                                                                                                                                '(' + currency + ')$'
        # net remit
    )

    international_summary = re.compile(
        '^ INTERNATIONAL FARES[ ]+'
        '(' + currency + ')[ ]+'  # credit
                         '(' + currency + ')[ ]+'  # cash
                                          '(' + currency + ')[ ]+'  # total
                                                           'TOTAL CANCELLATION PENALTY[ ]+'
                                                           '(' + currency + ')$'  # penalty
    )

    def __init__(self, report, airline, country):
        self.report = report
        self.airline = airline
        self.country = country
        self.totals = {
            'cc': 0.00,
            'ca': 0.00,
            'transaction_amount': 0.00,
            'fare_amount': 0.00,
            'std_comm': 0.00,
            'tax': 0.00,
            'pen': 0.00,

        }

    def process_value(self, value):
        for func in _process_funcs:
            m = func.regexp.match(value)
            if m:
                func(self, *m.groups())
                break
        else:
            pass

    @register(transaction_header)
    def process_transaction_header(self, ag_val1, ag_val2, ag_val3, trade_name, batch_code, date):

        self.agency, created = Agency.objects.get_or_create(
            agency_no=(ag_val1 + ag_val2 + ag_val3).zfill(8), country=self.country,
            defaults={
                'trade_name': trade_name
            },
        )
        self.date = datetime.datetime.strptime(date, "%m/%d/%y").date()

    @register(ticket_header)
    def process_ticket_header(self, card_no, ticket_no, processed_date, international, charge, charge_ex_tax, commission, rate, tax,
                              tax_type):

        transaction_values = {
            'card_code': card_no,
            'ticket_no': ticket_no,
            'transaction_type': 'TKT',
            'transaction_amount': get_float(charge),
            'fare_amount': get_float(charge_ex_tax),
            'cobl_amount': get_float(charge_ex_tax),
            'international': bool(international),
            'std_comm_rate': get_float(rate),
            'std_comm_amount': get_float(commission),
            # 'issue_date': self.date,

        }

        try:
            billdate = datetime.datetime.strptime(processed_date[0:2] +"-" +processed_date[3:]+ "-" + str(self.date.year), '%m-%d-%Y').date()
            transaction_values['issue_date'] = billdate
        except:
            transaction_values['issue_date'] = self.date

        self.transaction = self.create_transaction(transaction_values)

        tax_values = {
            'amount': get_float(tax),
            'type': tax_type,
        }


        self.tax = self.create_tax(tax_values)

    @register(tax_id)
    def process_tax_id(self, tax, tax_type):
        tax_values = {
            'amount': get_float(tax),
            'type': tax_type,
        }
        self.tax = self.create_tax(tax_values)

    @register(even_exchange_transaction)
    def process_even_exchange_transaction(
        self, ticket_no, charge, commission, tax, tax_type):
        transaction_values = {
            'card_code': '',
            'ticket_no': ticket_no,
            'transaction_type': 'EVEN EXCH',
            'fare_amount': get_float(charge),
            'cobl_amount': get_float(charge),
            'std_comm_amount': get_float(commission),
            'issue_date': self.date,
        }
        self.transaction = self.create_transaction(transaction_values)

        tax_values = {
            'amount': get_float(tax),
            'type': tax_type,
        }


        self.tax = self.create_tax(tax_values)

    @register(refund_or_exchange_transaction)
    def process_refund_or_exchange_transaction(self, card_no, ticket_no, date1, date2, international, total, ex_tax_val,
                                               commission, tax, tax_type):

        transaction_values = {
            'card_code': card_no,
            'ticket_no': ticket_no,
            'transaction_type': 'RFND',
            'transaction_amount': get_float(total),
            'fare_amount': get_float(ex_tax_val),
            'cobl_amount': get_float(ex_tax_val),
            'std_comm_amount': get_float(commission),
            'issue_date': self.date,
        }
        self.transaction = self.create_transaction(transaction_values)

        tax_values = {
            'amount': get_float(tax),
            'type': tax_type,
        }


        self.tax = self.create_tax(tax_values)

    @register(credit_debit_note_transaction)
    def process_credit_debit_note_transaction(self, ticket_no, amount):

        transaction_values = {
            'card_code': '',
            'ticket_no': ticket_no,
            'transaction_type': 'ACM' if get_float(amount) < 0 else 'ADM',
            'transaction_amount': get_float(amount),
            'fare_amount': get_float(amount),
            'cobl_amount': get_float(amount),
            'issue_date': self.date,

        }
        self.transaction = self.create_transaction(transaction_values)

    @register(exchange_transaction)
    def process_exchange_transaction(self, code, ticket_no, penalty, test):
        penalty = get_float(penalty)

        if penalty and test:
            transaction_values = {
                'card_code': code,
                'ticket_no': ticket_no,
                'transaction_type': 'EXCH',
                'pen': penalty,
                'pen_type': 'CANCEL PEN',
                'issue_date': self.date,
            }
        else:
            transaction_values = {
                'card_code': code,
                'ticket_no': ticket_no,
                'transaction_type': 'EXCH',
                'issue_date': self.date,
            }

        self.transaction = self.create_transaction(transaction_values)


    @register(refund_transaction)
    def process_refund_transaction(self, code, identifier, ticket_no, penalty, test):
        penalty = get_float(penalty)
        if penalty and test:
            transaction_values = {
                'ticket_no': ticket_no,
                'transaction_type': identifier,
                'issue_date': self.date,
                'amount': penalty,
                'type': 'CP'
            }
        else:
            transaction_values = {
                'ticket_no': ticket_no,
                'transaction_type': identifier,
                'issue_date': self.date,
            }

        self.transaction = self.create_tax(transaction_values)


    @register(AAD_transaction)
    def process_AAD_transaction(self, id_no, ticket_no, commission):

        transaction_values = {
            'card_code': '',
            'ticket_no': ticket_no,
            'transaction_type': 'AAD',
            'fare_amount': -1 * get_float(commission),
            'cobl_amount': -1 * get_float(commission),
            'issue_date': self.date,

        }
        self.transaction = self.create_transaction(transaction_values)

    @register(void_header)
    def process_void_header(self, ag_val1, ag_val2, ag_val3, batch_code, date):

        self.agency, created = Agency.objects.get_or_create(
            agency_no=int(ag_val1 + ag_val2 + ag_val3), country=self.country,
        )
        self.date = datetime.datetime.strptime(date, "%m/%d/%y").date()

    @register(void_ticket)
    def process_void_ticket(self, card_no, ticket_no, amount):
        transaction_values = {
            'card_code': card_no,
            'ticket_no': ticket_no,
            'transaction_type': 'VOID',
            'transaction_amount': None,
            'issue_date': self.date,

        }
        self.transaction = self.create_transaction(transaction_values)

    @register(can_void_ticket)
    def process_void_ticket(self, ticket_no):
        transaction_values = {
            'card_code': '',
            'ticket_no': ticket_no,
            'transaction_type': 'VOID',
            'issue_date': self.date,
            'transaction_amount': None,
            'fare_amount': 0.00
        }
        self.transaction = self.create_transaction(transaction_values)

    @register(transaction_summary)
    def process_transaction_summary(self, credit, cash, total, fare, comm, tax, adjust, ccrn_comm, net_remit):
        self.totals['cc'] = self.totals['cc'] + get_float(credit)
        self.totals['ca'] = self.totals['ca'] + get_float(cash)

    @register(international_summary)
    def process_international_summary(self, credit, cash, total, penalty):
        self.totals['pen'] = self.totals['pen'] + get_float(penalty)

    # Common functions

    def create_transaction(self, transaction_details_values):
        transaction = Transaction.objects.create(
            agency=self.agency,
            report=self.report,
            ticket_no=transaction_details_values.get('ticket_no'),
            transaction_type=transaction_details_values.get('transaction_type', None),
            issue_date=transaction_details_values.get('issue_date', self.date),
            transaction_amount = transaction_details_values.get('transaction_amount', 0.00),
            fare_amount = transaction_details_values.get('fare_amount', 0.00),
            card_code = transaction_details_values.get('card_code', None),
            international = transaction_details_values.get('international', False),
            # pen = transaction_details_values.get('pen', None),
            # pen_type = transaction_details_values.get('pen_type', None),
            cobl_amount = transaction_details_values.get('cobl_amount', 0.00),
            std_comm_rate = transaction_details_values.get('std_comm_rate', 0.00),
            std_comm_amount = transaction_details_values.get('std_comm_amount', 0.00),
        )
        pen = transaction_details_values.get('pen', None)
        pen_type = transaction_details_values.get('pen_type', None)
        if pen_type == 'CANCEL PEN':
            tax_values = {
            'amount': pen,
            'type': 'CP'
        }
            self.tax = self.create_tax(tax_values)
        return transaction

    def create_tax(self, values):
        
        tax_type = values.get('type', None)
        amount = values.get('amount', 0.00)
        
        if tax_type:
            if tax_type in ['CP','MF','YQ', 'YR']:
                charge = Charges.objects.create(amount=amount, type=tax_type, transaction=self.transaction)
            else: 
                tax = Taxes.objects.create(amount=amount, type=tax_type,
                                        transaction=self.transaction)
        return True


def process_carrier_report(text_file, request=None):
    """Parse Carrier Report file"""
    filedata = open(text_file, 'r', encoding="utf-8")
    header_values = None
    cr_ped_values = None
    cr_airline_values = None

    if request:
        # country = Country.objects.get(id=request.session.get('country'))
        if request.POST.get("from_scheduler") is not None:
            country = Country.objects.get(id=request.POST.get("countrycode"))
        else:
            country = Country.objects.get(id=request.session.get('country'))
    else:
        country = None
    logging.info("process_carrier_report      "+str(country))
    try:
        line = filedata.readline()
        m = re.match(cr_header, line)
        if m:
            header_values = m.groupdict()
        line = filedata.readline()
        m = re.match(cr_ped, line)
        if m:
            cr_ped_values = m.groupdict()
        line = filedata.readline()
        m = re.match(cr_airline, line)
        if m:
            cr_airline_values = m.groupdict()
        if header_values and cr_ped_values and cr_airline_values:
            try:
                airline = Airline.objects.get(code=cr_airline_values.get('code'), country=country)

            except  :
                logging.info("It seems like there is no airline with this 3 digit code.   => "+cr_airline_values.get('code'))
                return ("It seems like there is no airline with this 3 digit code.")
        else:
            logging.info("File does not seem to have the right format")
            return ("File does not seem to have the right format")
        end_date = datetime.datetime.strptime(cr_ped_values.get('ped'), "%m/%d/%y").date()

        try:
            report_period = ReportPeriod.objects.get(ped=end_date, country=country)
        except  :
            logging.info('PED is not found')
            return 'PED is not found'

        rf, created = ReportFile.objects.update_or_create(
            report_period=report_period, airline=airline, country=country, defaults={
                'ref_no': header_values.get('ref_no').strip(),
                'file': text_file.split('media/')[-1],
            })

        Transaction.objects.filter(report=rf).delete()

        commission_history = CommissionHistory.objects.filter(airline=airline, type='M',
                                                              from_date__lte=end_date)
        if commission_history.exists():
            allowed_commission_rate = commission_history.order_by('-from_date').first().rate

        process_data = ProcessData(rf, airline, country)
        # process remainder of file
        endline = ''
        for line in filedata.readlines():
            process_data.process_value(line)
            endline = line

        trans_data = Transaction.objects.exclude(transaction_type__startswith='SP',
                                                 agency__agency_no='6999001').filter(report=rf).aggregate(
            transaction_amount=Sum('transaction_amount'),
            fare_amount=Sum('fare_amount'),
            std_comm=Sum('std_comm_amount'),
            cobl_amount=Sum('cobl_amount'),
            # cc=Sum('cc'),
            # ca=Sum('ca')
        )
        tax_tot = Taxes.objects.filter(transaction__report=rf).aggregate(total_tax=Sum('amount')).get('total_tax')
        tot_charg = Charges.objects.filter(transaction__report=rf).aggregate(total_charges=Sum('amount')).get(
            'total_charges')

        # process_data.totals['cc'] = self.totals['cc'] + get_float(credit)
        # process_data.totals['ca'] = self.totals['ca'] + get_float(cash)
        process_data.totals['transaction_amount'] = trans_data.get('transaction_amount')
        process_data.totals['fare_amount'] = trans_data.get('fare_amount')
        process_data.totals['std_comm'] = trans_data.get('std_comm')
        process_data.totals['tax'] = tax_tot if tax_tot else 0.00
        process_data.totals['fandc'] = tot_charg if tot_charg else 0.00

        ReportFile.objects.filter(id=rf.id).update(**process_data.totals)

    except Exception as e:
        print("Error Occured : ", e)
        logging.info("process_carrier_report             ========"+e)
        pass


filename_test = re.compile('DISBADV\d{6}[a|b]0?[_| ]([a-zA-Z0-9]{2})')

line1 = re.compile(
    "^[ ]+REPORT ID - \w+-\w+[ ]+"
    "AIRLINES REPORTING CORPORATION[ ]+"
    "REF NBR - [\d-]+$")
line2 = re.compile(
    "^[ ]+PAGE -[ ]+\d+[ ]+"
    "CARRIER DISBURSEMENT ADVICE[ ]+"
    "CUR PED - (\d{2}-\d{2}-\d{2})$")  # filedate
line3 = re.compile(
    "^[ ]+RUN DATE[ ]+-[ ]+"
    "(\d{2}/\d{2}/\d{2})$")  # rundate
line4 = re.compile(
    "^[ ]+AIRLINE:[ ]+"
    "(\d+)-\d+[ ]+.+$")  # airline code


def process_disbursement_advice(text_file, filename, request=None):
    """Parse Disbursement Advice file"""
    filedata = open(text_file, 'r', encoding="utf-8")
    if request:
        if request.POST.get("from_scheduler") is not None:
            country = Country.objects.get(id=request.POST.get("countrycode"))
        else:
            country = Country.objects.get(id=request.session.get('country'))
    else:
        country = None
    logging.info("process_disbursement_advice     "+str(country))
    airline_test = filename_test.match(filename)
    if not airline_test:
        logging.info("Bad Disbursment filename")
        return 'Bad Disbursment filename'
    try:
        ai = airline_test.group(1)
    except:
        logging.info("No such airline in file name")
        return 'No such airline in file name'

    line = filedata.readline()

    if not line1.match(line):
        logging.info("Incorrect File format")
        return "Incorrect File format"

    disbursement = None
    arc = None
    bankpay = {}
    rundate = None
    filedate = None
    pending_payments = False
    for i in filedata.readlines():

        m = line2.match(i)
        if m:
            (filedate,) = m.groups()
            filedate = datetime.datetime.strptime(filedate, "%m-%d-%y")
            continue

        m = line3.match(i)
        if m:
            (rundate,) = m.groups()
            rundate = datetime.datetime.strptime(rundate, "%m/%d/%y")
            continue

        m = line4.match(i)
        if m:
            (airl_code,) = m.groups()

            try:

                airline = Airline.objects.get(code=airl_code, country=country)
            except  :
                logging.info("It seems like there is no airline with this 3 digit code. => "+airl_code)
                return ("It seems like there is no airline with this 3 digit code.")

            defaults = {
                'filedate': filedate,
                'rundate1': rundate,
                'file1': text_file.split('media/')[-1],
                'arc_deduction': 0.0,
                'arc_fees': 0.0,
                'arc_tot': 0.0,
                'arc_reversal': 0.0,
            }

            try:
                report_period = ReportPeriod.objects.filter(ped__gte=filedate, country=country).order_by('ped').first()
                if not report_period:
                    return 'PED is not found'
            except  :
                return 'PED is not found'
            try:
                rf = ReportFile.objects.get(report_period=report_period, airline=airline, country=country)
            except  :
                return "Upload Carrier Report First"

            disbursement, created = Disbursement.objects.get_or_create(report_period=report_period, airline=airline,
                                                                       defaults=defaults)
            if not created:
                error = disbursement.is_filed(rundate, text_file)
                if error:
                    return error
        else:
            logging.info("REGEX DIDN'T MATCH")
    disbursement.save()
    disbursement.reprocess_files()


ref_line = re.compile(
    "AIRLINES REPORTING CORPORATION[ ]+"
    "REF NBR -[ ]+[\d-]+$")
filedate_line = re.compile(
    "ITEMIZATION OF CARRIER DEDUCTIONS[ ]+"
    "CUR PED -[ ]+(\d{2}/\d{2}/\d{2})$")  # filedate
carrier_line = re.compile(
    "^CARRIER NAME:[ ]+"
    "(\d+)[ ]+.+$")  # airline code
type_map = {
    'LESS PROCESSING': "PROCESSING",
    'SS / OH': "SERVICES / OVERHEAD",
    'ADJUSTMENTS': "ADJUSTMENTS",
    'GROSS DISBURSEMENTS': "GROSS DISBURSEMENTS",
    'NET DISBURSEMENTS': "NET DISBURSEMENTS"
}
types_regexp = "|".join("(?:%s)" % t for t in type_map.keys())
deduction_line = re.compile(
    "^[ ]+(" + types_regexp + "):[ ]*"
                              "(" + currency + ")?[ ]*(NA)$")

deduction_line_gross = re.compile(r'^BILLING SUMMARY:\s+(GROSS DISBURSEMENTS):\s+(\d*,?\d*,?\d*\.?\d+-?)\s+(NA)$')


def process_carrier_deductions(text_file, filename, request=None):
    """Parse Disbursement Advice file"""

    filedata = open(text_file, 'r', encoding="utf-8")
    if request:
        if request.POST.get("from_scheduler") is not None:
            country = Country.objects.get(id=request.POST.get("countrycode"))
        else:
            country = Country.objects.get(id=request.session.get('country'))
    else:
        country = None
    logging.info("process_carrier_deductions     "+str(country))
    line = filedata.readline()

    if not ref_line.match(line):
        logging.info("Incorrect File format")
        return "Incorrect File format"

    for line in filedata.readlines():
        m = filedate_line.match(line)
        if m:
            (filedate,) = m.groups()
            filedate = datetime.datetime.strptime(filedate, "%m/%d/%y")
            continue

        m = carrier_line.match(line)
        if m:
            (carrier_code,) = m.groups()
            try:
                airline = Airline.objects.get(code=carrier_code, country=country)
            except  :
                logging.info("It seems like there is no airline with this 3 digit code. => "+carrier_code)
                return ("It seems like there is no airline with this 3 digit code.")

            defaults = {
                'file': text_file.split('media/')[-1],
                'filedate': filedate,
            }

            try:
                report_period = ReportPeriod.objects.filter(ped__gte=filedate, country=country).order_by('ped').first()
                if not report_period:
                    logging.info("PED is not found")
                    return 'PED is not found'
            except  :
                return 'PED is not found'

            try:
                rf = ReportFile.objects.get(report_period=report_period, airline=airline, country=country)
            except  :
                logging.info("Upload Carrier Report First")
                return "Upload Carrier Report First"

            deductions, created = CarrierDeductions.objects.get_or_create(report_period=report_period, airline=airline,
                                                                          defaults=defaults)
            if not created:
                logging.info("This report has already been imported.")
                return 'This report has already been imported.'
            continue
        else:
            logging.info("REGEX DIDN'T MATCH")
        mg = deduction_line_gross.match(line)
        if mg:
            (type, amount, pending) = mg.groups()
            d = Deduction(
                report=deductions,
                type=type_map[type],
                amount=get_float(amount),
                pending=pending == 'NA')
            d.save()
            continue

        m = deduction_line.match(line)
        if m:
            (type, amount, pending) = m.groups()
            d = Deduction(
                report=deductions,
                type=type_map[type],
                amount=get_float(amount),
                pending=pending == 'NA')
            d.save()
            continue

    deductions.save()
    # return deductions


def remittance(text_file):
    """Parse billing details file"""
    # / home / fingent / Documents / CA_BSP_2018_Agents_Calendar.txt
    regex = re.compile(
        r"\s+\w+\s+\w+-(?P<from>\d{2}-\w{3}-\w{2})\s+\w+-(?P<to>\d{2}-\w{3}-\w{2})\s+\w+-(?P<nope>\d{2}-\w{3}-\w{2})\s+\w+-(?P<remittance>\d{2}-\w{3}-\w{2})")

    filedata = open(text_file, 'r', encoding="utf-8")
    reader = enumerate(filedata.readlines())
    try:
        while reader:
            number, line = next(reader, None)
            m = re.match(regex, line)
            if m:
                values = m.groupdict()
                remi, created = Remittance.objects.get_or_create(
                    ped=datetime.datetime.strptime(values.get('to'), "%d-%b-%y").date(),
                    defaults={'remittance': datetime.datetime.strptime(values.get('remittance'), "%d-%b-%y").date()})
    except Exception as e:
        print("Error Occured : ", e)


@shared_task
def re_process(airline, start, end, task_id):
    task = ReprocessFile.objects.get(pk=task_id)
    try:
        start = datetime.datetime.strptime(start, '%d %B %Y')
        end = datetime.datetime.strptime(end, '%d %B %Y')
        all_files = ReportFile.objects.filter(report_period__ped__range=[start, end])
        if airline:
            all_files = all_files.filter(airline=airline)
        lst = []
        for file in all_files:
            aa = process_billing_details(None, None, file)
            lst.append(aa)
        task.message = str(lst)
        task.is_done = True
        task.status = 'success'
        task.save()
    except:
        task.is_done = True
        task.status = 'failed'
        task.save()
    return True




def process_excelfile(file, request=None):
    store_list = []
    months = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October',
              'November', 'December']
    weeks = ['Week 1', 'Week 2', 'Week 3', 'Week 4', 'Week 5', 'Week 6']

    country = Country.objects.get(id=request.session.get('country'))
    wb = load_workbook(file)

    if len(wb.worksheets) > 1:
        return ("Calendar file contains more than one sheet")

    for sheet in wb.worksheets:
        cnt = 0
        todateList = []
        myDict = {}
        # iterating over the rows and
        # getting value from each cell in row
        reader = sheet.iter_rows()

        for row in reader:
            values = [cell.value for cell in row]

            if cnt == 0:
                # checking data for year in sheet

                if checkYear(values[0]) != 'Y':
                    return checkYear(values[0])

                year = int(values[0])
                if ReportPeriod.objects.filter(year=int(values[0]), country=request.session.get('country')).exists():
                    return ("Data already present for the year " + str(year))

            elif cnt == 1:
                fromdatefield = values[2]
                todatefield = values[3]
                lastfield = values[4]
                # import pdb;pdb.set_trace()

                if values[0].lower() == 'month' and values[
                    1].lower() == 'week' and fromdatefield.strip().lower() == 'from date' and todatefield.strip().lower() == 'to date':
                    # checking file with selected country

                    if (lastfield.strip().lower() == 'remittance date' and country.name == 'United States') or (
                        lastfield.strip().lower() == 'disbursement date' and country.name != 'United States'):
                        return ("Selected a wrong calendar file")
                    else:
                        pass
                else:
                    return ("In calendar file, cell headings are not proper")

            else:
                del values[5:len(values)]

                if None in values and set(values) != set([None]):
                    return ("Some cell values were missing in the calendar file")
                elif None in values and set(values) == set([None]):
                    pass
                else:
                    if values[0] not in months:
                        return ("Month specified in calendar file is incorrect")
                    if values[1] not in weeks:
                        return ("Week specified in calendar file is incorrect")

                    # Week duplication check in calendar
                    if values[0] not in myDict:
                        myDict[values[0]] = []

                    if values[1] in myDict[values[0]]:
                        return ('Some weeks are duplicated in the uploaded calendar file')
                    else:
                        myDict[values[0]].append(values[1])

                    month_num = int(months.index(values[0])) + 1

                    # date format check for from date, to date and remittance date fields
                    fromdatefieldvalidate = validateDateFormat(values[2], fromdatefield, year, month_num)
                    todatefieldvalidate = validateDateFormat(values[3], todatefield, year, month_num)
                    remdatefieldvalidate = validateDateFormat(values[4], lastfield, year, month_num)

                    if fromdatefieldvalidate != 'Y':
                        return (fromdatefieldvalidate)

                    if todatefieldvalidate != 'Y':
                        return (todatefieldvalidate)

                    if remdatefieldvalidate != 'Y':
                        return (remdatefieldvalidate)

                    # compare from date with to date
                    compDate = compareDate(values[2], values[3])

                    if compDate == 'N':
                        return ("'" + todatefield + "' must be greater than '" + fromdatefield + "'")
                    elif compDate != 'Y':
                        return compDate
                    else:
                        pass

                    if todateList:
                        compprevtoDate = compareDate(todateList[-1], values[2])

                        if compprevtoDate == 'N':
                            return ("'" + fromdatefield + "'  must be greater than previous '" + todatefield + "'")
                        elif compDate != 'Y':
                            return compDate
                        else:
                            pass

                    todateList.append(values[3])

                    store_list.append(ReportPeriod(year=year,
                                                   month=months.index(values[0]) + 1,
                                                   week=values[1].replace("Week ", ""),
                                                   ped=datetime.datetime.strptime(
                                                       str(values[3]).replace(" 00:00:00", ""), '%Y-%m-%d'),
                                                   from_date=datetime.datetime.strptime(
                                                       str(values[2]).replace(" 00:00:00", ""), '%Y-%m-%d'),
                                                   remittance_date=datetime.datetime.strptime(
                                                       str(values[4]).replace(" 00:00:00", ""), '%Y-%m-%d'),
                                                   country=country))

            cnt = cnt + 1

        if len(store_list) > 0 and len(store_list) < 100:
            try:
                ReportPeriod.objects.bulk_create(store_list)
            except :
                return ("Calendar file has some issue")
        elif len(store_list) >= 100:
            return ("Calendar file contains excess data")
        else:
            return ("Calendar file is empty")

    return ("success")

# def process_excelfile(file, request=None):
#     store_list = []
#     months = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October',
#               'November', 'December']
#     weeks = ['Week 1', 'Week 2', 'Week 3', 'Week 4', 'Week 5', 'Week 6', 'Week 7', 'Week 8', 'Week 9', 'Week 10']

#     country = Country.objects.get(id=request.session.get('country'))
#     # ReportPeriod.objects.filter(country = country).delete()
#     wb = load_workbook(file)

#     if len(wb.worksheets) > 1:
#         return ("Calendar file contains more than one sheet")

#     for sheet in wb.worksheets:
#         cnt = 0
#         todateList = []
#         myDict = {}
#         # iterating over the rows and
#         # getting value from each cell in row
#         reader = sheet.iter_rows()

#         for row in reader:
#             values = [cell.value for cell in row]

#             if cnt == 0:
#                 # checking data for year in sheet

#                 if checkYear(values[0]) != 'Y':
#                     return checkYear(values[0])

#                 year = int(values[0])
#                 if ReportPeriod.objects.filter(year=int(values[0]), country=request.session.get('country')).exists():
#                     return ("Data already present for the year " + str(year))

#             elif cnt == 1:
#                 fromdatefield = values[2]
#                 todatefield = values[3]
#                 lastfield = values[4]
#                 # import pdb;pdb.set_trace()

#                 if values[0].lower().strip() == 'month' and values[
#                     1].lower().strip() == 'week' and fromdatefield.strip().lower() == 'from date' and todatefield.strip().lower() == 'to date':
#                     # checking file with selected country

#                     if (lastfield.strip().lower() == 'remittance date' and country.name == 'United States') or (
#                         lastfield.strip().lower() == 'disbursement date' and country.name != 'United States'):
#                         return ("Selected a wrong calendar file")
#                     else:
#                         pass
#                 else:
#                     return ("In calendar file, cell headings are not proper")

#             else:
#                 del values[5:len(values)]

#                 if None in values and set(values) != set([None]):
#                     return ("Some cell values were missing in the calendar file")
#                 elif None in values and set(values) == set([None]):
#                     pass
#                 else:
#                     if values[0] not in months:
#                         return ("Month specified in calendar file is incorrect")
#                     if values[1] not in weeks:
#                         return ("Week specified in calendar file is incorrect")

#                     # Week duplication check in calendar
#                     if values[0] not in myDict:
#                         myDict[values[0]] = []

#                     if values[1] in myDict[values[0]]:
#                         return ('Some weeks are duplicated in the uploaded calendar file')
#                     else:
#                         myDict[values[0]].append(values[1])

#                     month_num = int(months.index(values[0])) + 1

#                     # date format check for from date, to date and remittance date fields
#                     fromdatefieldvalidate = validateDateFormat(values[2], fromdatefield, year, month_num)
#                     todatefieldvalidate = validateDateFormat(values[3], todatefield, year, month_num)
#                     remdatefieldvalidate = validateDateFormat(values[4], lastfield, year, month_num)

#                     if fromdatefieldvalidate != 'Y':
#                         return (fromdatefieldvalidate)

#                     if todatefieldvalidate != 'Y':
#                         return (todatefieldvalidate)

#                     if remdatefieldvalidate != 'Y':
#                         return (remdatefieldvalidate)

#                     # compare from date with to date
#                     compDate = compareDate(values[2], values[3])

#                     if compDate == 'N':
#                         return ("'" + todatefield + "' must be greater than.. '" + fromdatefield + "'")
#                     elif compDate != 'Y':
#                         return compDate
#                     else:
#                         pass

#                     if todateList:
#                         # compprevtoDate = compareDate(todateList[-1], values[2])

#                         # if compprevtoDate == 'N':
#                         #     return ("'" + fromdatefield + "'  must be greater than previous...# '" + todatefield + "'")
#                         # elif compDate != 'Y':
#                         #     return compDate
#                         # else:
#                         pass

#                     todateList.append(values[3])

#                     store_list.append(ReportPeriod(year=year,
#                                                    month=months.index(values[0]) + 1,
#                                                    week=values[1].replace("Week ", ""),
#                                                    ped=datetime.datetime.strptime(
#                                                        str(values[3]).replace(" 00:00:00", ""), '%Y-%m-%d'),
#                                                    from_date=datetime.datetime.strptime(
#                                                        str(values[2]).replace(" 00:00:00", ""), '%Y-%m-%d'),
#                                                    remittance_date=datetime.datetime.strptime(
#                                                        str(values[4]).replace(" 00:00:00", ""), '%Y-%m-%d'),
#                                                    country=country))

#             cnt = cnt + 1
#             print(store_list,"//////////")
#         # if len(store_list) > 0 and len(store_list) < 100:
#         try:
#             ReportPeriod.objects.bulk_create(store_list)
#         except :
#             return ("Calendar file has some issue//////////////////////////////////")
#         # elif len(store_list) >= 100:
#         #     return ("Calendar file contains excess data/////////////////////////////////")
#         # else:
#         #     return ("Calendar file is empty/////////////////////////////////")

#     return ("success")


def checkYear(year):
    if (isinstance(year, int) or isinstance(year, float)):
        if len(str(abs(int(year)))) != 4:
            return ("Year specified in calendar file is incorrect")
        else:
            return ("Y")
    else:
        return ("Year specified in calendar file is incorrect")


def validateDateFormat(date, field, year, month):
    try:
        if type(date) is not datetime.datetime:
            return ("Incorrect '" + field.capitalize() + "' date format")
    except ValueError:
        return ("Incorrect '" + field.capitalize() + "' date format")
    return ('Y')


def compareDate(fromDate, toDate):
    try:
        date1 = fromDate.date()
        date2 = toDate.date()

        if date2 <= date1:
            return ('N')
    except ValueError:
        return ('Issue with date format')

    return ('Y')


def num_quantize(value, n_point=2):
    """
    :param value:
    :param n_point:
    :return:
    """
    from decimal import localcontext, Decimal, ROUND_HALF_UP
    with localcontext() as ctx:
        ctx.rounding = ROUND_HALF_UP
        if value:
            d_places = Decimal(10) ** -n_point
            # Round to two places
            value = Decimal(value).quantize(d_places)
        return value
